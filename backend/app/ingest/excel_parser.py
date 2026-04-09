"""Excel ingest pipeline — parse, dedup, classify, sort.

Pure function: Excel bytes in → ParsedBatch out. No DB writes.
"""
from __future__ import annotations

import uuid
from dataclasses import dataclass, field
from datetime import datetime
from io import BytesIO

from openpyxl import load_workbook

NULL_VALUES = {"NULL", "null", "", None}

# Structured columns mapped from Excel header → Case model field
COLUMN_MAP = {
    "ExamId": "exam_id",
    "FirstName": "first_name",
    "LastName": "last_name",
    "dob": "dob",
    "policynum": "policy_num",
    "PatientZipCode": "patient_zip",
    "CenterNPI": "center_npi",
    "CenterAbbr": "center_abbr",
    "cptcode": "cpt_code",
    "ICD1": "icd1",
    "ICD2": "icd2",
    "ICD3": "icd3",
    "ICD4": "icd4",
    "ICD5": "icd5",
    "ReferringNPI": "referring_npi",
    "CarrierId": "carrier_id",
    "FileKey": "file_key",
    "AttachmentId": "attachment_id",
    "IsStat": "is_stat",
    "ScheduledDateTime": "scheduled_dt",
    "AuthExamId": "auth_exam_id",
    "OrderRequestID": "order_request_id",
}

# Required headers — must be present in the first row
REQUIRED_HEADERS = {"ExamId", "FirstName", "LastName", "dob", "policynum", "cptcode", "CenterNPI"}


@dataclass
class ParsedBatch:
    filename: str
    total_rows: int
    duplicate_rows: int
    unique_cases: int
    stat_count: int
    hold_count: int
    cases: list[dict] = field(default_factory=list)


def parse_excel(file_bytes: bytes, filename: str) -> ParsedBatch:
    """Parse Carelon Excel file → ParsedBatch with sorted case dicts."""
    wb = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(min_row=1, values_only=False))
    if not rows:
        wb.close()
        raise ValueError("Upload failed: Excel file is empty.")

    # First row must be headers
    headers = [str(cell.value).strip() if cell.value is not None else None for cell in rows[0]]
    found = {h for h in headers if h is not None}
    missing = REQUIRED_HEADERS - found
    if missing:
        wb.close()
        raise ValueError(f"Upload failed: missing required headers: {', '.join(sorted(missing))}")

    data_rows = rows[1:]
    total_rows = len(data_rows)

    # Convert to list of dicts
    raw_records = []
    for row in data_rows:
        record = {}
        for i, header in enumerate(headers):
            if header is not None:
                val = row[i].value if i < len(row) else None
                record[header] = val
        raw_records.append(record)

    wb.close()

    # Dedup on ExamId — first occurrence wins
    seen_exam_ids: set[str] = set()
    unique_records: list[dict] = []
    for record in raw_records:
        exam_id = _clean_id(record.get("ExamId"))
        if exam_id and exam_id not in seen_exam_ids:
            seen_exam_ids.add(exam_id)
            unique_records.append(record)

    duplicate_rows = total_rows - len(unique_records)

    # Map and classify each record
    cases: list[dict] = []
    stat_count = 0
    hold_count = 0

    for record in unique_records:
        case_dict = _map_record(record)
        _classify(case_dict)

        if case_dict["state"] == "PENDING_STAT":
            stat_count += 1
        elif case_dict["state"] == "HOLD":
            hold_count += 1

        cases.append(case_dict)

    # Sort: priority → scheduled_dt → ingested order
    cases.sort(key=_sort_key)

    return ParsedBatch(
        filename=filename,
        total_rows=total_rows,
        duplicate_rows=duplicate_rows,
        unique_cases=len(cases),
        stat_count=stat_count,
        hold_count=hold_count,
        cases=cases,
    )


def _map_record(record: dict) -> dict:
    """Split Excel row into structured fields + raw_data JSONB."""
    case = {"id": str(uuid.uuid4())}

    # Map structured fields
    for excel_col, model_field in COLUMN_MAP.items():
        raw_val = record.get(excel_col)

        if model_field == "is_stat":
            case[model_field] = _clean_bool(raw_val)
        elif model_field == "scheduled_dt":
            case[model_field] = _clean_datetime(raw_val)
        elif model_field in ("first_name", "last_name"):
            case[model_field] = _clean_name(raw_val)
        elif model_field == "dob":
            case[model_field] = _clean_dob(raw_val)
        elif model_field in ("exam_id", "center_npi", "carrier_id",
                             "referring_npi", "attachment_id",
                             "cpt_code", "order_request_id", "auth_exam_id"):
            case[model_field] = _clean_id(raw_val)
        elif model_field in ("icd1", "icd2", "icd3", "icd4", "icd5"):
            case[model_field] = _clean_nullable_str(raw_val)
        else:
            case[model_field] = _clean_nullable_str(raw_val)

    # Everything goes into raw_data (full original row)
    raw_data = {}
    for key, val in record.items():
        if val in NULL_VALUES:
            raw_data[key] = None
        else:
            # Serialize datetime for JSON
            if isinstance(val, datetime):
                raw_data[key] = val.isoformat()
            else:
                raw_data[key] = val

    # Data quality flags
    flags = []
    if not case.get("icd1"):
        flags.append("no_icd")
    if not case.get("referring_npi"):
        flags.append("no_referring")
    if flags:
        raw_data["_flags"] = flags

    case["raw_data"] = raw_data
    return case


def _classify(case: dict) -> None:
    """Assign state and sort_priority."""
    has_member_search = all([
        case.get("first_name"),
        case.get("last_name"),
        case.get("dob"),
        case.get("policy_num"),
        case.get("center_npi"),
    ])

    if not has_member_search:
        missing = []
        for field_name in ("first_name", "last_name", "dob", "policy_num", "center_npi"):
            if not case.get(field_name):
                missing.append(field_name)
        case["state"] = "HOLD"
        case["sort_priority"] = 10
        case["hold_reason"] = f"Missing required fields: {', '.join(missing)}"
    elif case.get("is_stat"):
        case["state"] = "PENDING_STAT"
        case["sort_priority"] = 1
    else:
        case["state"] = "PENDING_NOTES"
        # Has RIS attachment if attachment_id is present and not "0"
        has_ris = case.get("attachment_id") and case["attachment_id"] != "0"
        case["sort_priority"] = 2 if has_ris else 3


def _sort_key(case: dict):
    """Sort: priority ASC → scheduled_dt ASC (nulls last) → id for stability."""
    priority = case.get("sort_priority", 99)
    sched = case.get("scheduled_dt")
    # None sorts last
    sched_key = sched if sched is not None else datetime.max
    return (priority, sched_key, case.get("id", ""))


# --- Cleaning helpers ---


def _clean_id(val) -> str | None:
    """All IDs as string — no float surprises."""
    if val in NULL_VALUES:
        return None
    s = str(val).strip()
    # Float from pandas/openpyxl: "1234567.0" → "1234567"
    if s.endswith(".0"):
        s = s[:-2]
    return s if s else None


def _clean_name(val) -> str | None:
    """Normalize to Title Case."""
    if val in NULL_VALUES:
        return None
    return str(val).strip().title()


def _clean_nullable_str(val) -> str | None:
    if val in NULL_VALUES:
        return None
    s = str(val).strip()
    return s if s else None


def _clean_dob(val) -> str | None:
    """Normalize DOB to MM/DD/YYYY format for portal submission."""
    if val in NULL_VALUES:
        return None
    if isinstance(val, datetime):
        return val.strftime("%m/%d/%Y")
    s = str(val).strip()
    if not s:
        return None
    # Try common formats
    from datetime import date as date_type
    import re
    # YYYY-MM-DD (ISO)
    m = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})$", s)
    if m:
        return f"{int(m.group(2)):02d}/{int(m.group(3)):02d}/{m.group(1)}"
    # MM/DD/YYYY (already correct)
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", s)
    if m:
        return f"{int(m.group(1)):02d}/{int(m.group(2)):02d}/{m.group(3)}"
    # M-D-YYYY or MM-DD-YYYY
    m = re.match(r"^(\d{1,2})-(\d{1,2})-(\d{4})$", s)
    if m:
        return f"{int(m.group(1)):02d}/{int(m.group(2)):02d}/{m.group(3)}"
    # Fallback: try dateutil
    try:
        from dateutil import parser as dtparse
        parsed = dtparse.parse(s)
        return parsed.strftime("%m/%d/%Y")
    except (ValueError, TypeError):
        return s  # Return as-is if unparseable


def _clean_bool(val) -> bool:
    if val in NULL_VALUES:
        return False
    return str(val).strip().upper() in ("YES", "TRUE", "1")


def _clean_datetime(val) -> datetime | None:
    if val in NULL_VALUES:
        return None
    if isinstance(val, datetime):
        return val
    try:
        from dateutil import parser as dtparse
        return dtparse.parse(str(val))
    except (ValueError, TypeError):
        return None
